"""
Complete working example: Generate w&k Connect CRM Report

Production-ready script for generating Excel reports with charts,
conditional formatting, and professional styling.

Usage:
    python crm_report_generator.py
"""

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from datetime import date, timedelta
import random


class CRMReportGenerator:
    """Professional CRM report generator for w&k Connect."""

    COMPANY_COLORS = {
        'header_bg': '059669',      # Green
        'header_text': 'FFFFFF',     # White
        'accent': '6366F1',          # Indigo
        'positive': '10B981',        # Green
        'negative': 'EF4444',        # Red
        'neutral': 'F59E0B'          # Orange
    }

    def __init__(self):
        """Initialize workbook."""
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def _create_title(self, title: str, row: int = 1):
        """Create report title with company branding."""
        self.ws.merge_cells(f'A{row}:F{row}')
        cell = self.ws[f'A{row}']
        cell.value = title
        cell.font = Font(name='Open Sans', size=16, bold=True, color=self.COMPANY_COLORS['header_bg'])
        cell.alignment = Alignment(horizontal='center', vertical='center')
        self.ws.row_dimensions[row].height = 30

    def _create_header_row(self, headers: list, row: int):
        """Create styled header row."""
        for col_idx, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color=self.COMPANY_COLORS['header_text'])
            cell.fill = PatternFill(
                start_color=self.COMPANY_COLORS['header_bg'],
                end_color=self.COMPANY_COLORS['header_bg'],
                fill_type='solid'
            )
            cell.alignment = Alignment(horizontal='center', vertical='center')

        self.ws.row_dimensions[row].height = 25

    def create_broker_performance_sheet(self, broker_data: list):
        """Create broker performance worksheet with chart."""
        self.ws.title = "Broker Performance"

        # Title
        self._create_title(f'w&k Connect CRM Report - {date.today().strftime("%B %Y")}', row=1)

        # Subtitle
        subtitle_row = 2
        self.ws.merge_cells(f'A{subtitle_row}:F{subtitle_row}')
        cell = self.ws[f'A{subtitle_row}']
        cell.value = f'Report Period: {(date.today() - timedelta(days=30)).strftime("%d.%m.%Y")} - {date.today().strftime("%d.%m.%Y")}'
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(size=10, color='6B7280')

        # Headers
        headers = ['Broker', 'Leads', 'Deals Closed', 'Revenue (€)', 'Conversion %', 'Status']
        self._create_header_row(headers, row=4)

        # Data rows
        start_row = 5
        for idx, broker in enumerate(broker_data, start=start_row):
            self.ws.cell(row=idx, column=1, value=broker['name'])
            self.ws.cell(row=idx, column=2, value=broker['leads'])
            self.ws.cell(row=idx, column=3, value=broker['deals'])

            # Revenue with currency format
            revenue_cell = self.ws.cell(row=idx, column=4, value=broker['revenue'])
            revenue_cell.number_format = '€#,##0.00'

            # Conversion rate with percentage format
            conversion_cell = self.ws.cell(row=idx, column=5, value=broker['conversion'])
            conversion_cell.number_format = '0.0%'

            # Status based on conversion rate
            status_cell = self.ws.cell(row=idx, column=6)
            if broker['conversion'] >= 0.15:
                status_cell.value = '✓ On Target'
                status_cell.font = Font(color=self.COMPANY_COLORS['positive'])
            elif broker['conversion'] >= 0.10:
                status_cell.value = '⚠ Below Target'
                status_cell.font = Font(color=self.COMPANY_COLORS['neutral'])
            else:
                status_cell.value = '✗ Needs Attention'
                status_cell.font = Font(color=self.COMPANY_COLORS['negative'])

        # Totals row
        total_row = len(broker_data) + start_row
        self.ws.cell(row=total_row, column=1, value='TOTAL')
        self.ws.cell(row=total_row, column=2, value=f'=SUM(B{start_row}:B{total_row-1})')
        self.ws.cell(row=total_row, column=3, value=f'=SUM(C{start_row}:C{total_row-1})')
        self.ws.cell(row=total_row, column=4, value=f'=SUM(D{start_row}:D{total_row-1})')
        self.ws.cell(row=total_row, column=4).number_format = '€#,##0.00'
        self.ws.cell(row=total_row, column=5, value=f'=AVERAGE(E{start_row}:E{total_row-1})')
        self.ws.cell(row=total_row, column=5).number_format = '0.0%'

        # Style totals row
        for col in range(1, 7):
            cell = self.ws.cell(row=total_row, column=col)
            cell.font = Font(bold=True, color=self.COMPANY_COLORS['header_bg'])

        # Add conditional formatting for conversion rates
        rule = ColorScaleRule(
            start_type='num', start_value=0, start_color=self.COMPANY_COLORS['negative'],
            mid_type='num', mid_value=0.15, mid_color=self.COMPANY_COLORS['neutral'],
            end_type='num', end_value=0.30, end_color=self.COMPANY_COLORS['positive']
        )
        self.ws.conditional_formatting.add(f'E{start_row}:E{total_row-1}', rule)

        # Add bar chart
        chart = BarChart()
        chart.title = "Broker Revenue Performance"
        chart.x_axis.title = "Broker"
        chart.y_axis.title = "Revenue (€)"
        chart.style = 11  # Professional style

        data = Reference(self.ws, min_col=4, min_row=4, max_row=total_row-1)
        cats = Reference(self.ws, min_col=1, min_row=start_row, max_row=total_row-1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        chart.height = 10  # cm
        chart.width = 20   # cm

        self.ws.add_chart(chart, f'H4')

        # Adjust column widths
        self.ws.column_dimensions['A'].width = 20
        self.ws.column_dimensions['B'].width = 12
        self.ws.column_dimensions['C'].width = 15
        self.ws.column_dimensions['D'].width = 15
        self.ws.column_dimensions['E'].width = 15
        self.ws.column_dimensions['F'].width = 18

    def create_monthly_trend_sheet(self, monthly_data: list):
        """Create monthly trend analysis sheet with line chart."""
        ws = self.wb.create_sheet("Monthly Trends")

        # Title
        ws.merge_cells('A1:E1')
        cell = ws['A1']
        cell.value = 'Monthly Performance Trends'
        cell.font = Font(size=14, bold=True, color=self.COMPANY_COLORS['header_bg'])
        cell.alignment = Alignment(horizontal='center')

        # Headers
        headers = ['Month', 'Total Leads', 'Deals Closed', 'Revenue (€)', 'Avg Conversion %']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color=self.COMPANY_COLORS['header_text'])
            cell.fill = PatternFill(
                start_color=self.COMPANY_COLORS['accent'],
                fill_type='solid'
            )
            cell.alignment = Alignment(horizontal='center')

        # Data
        start_row = 4
        for idx, month in enumerate(monthly_data, start=start_row):
            ws.cell(row=idx, column=1, value=month['month'])
            ws.cell(row=idx, column=2, value=month['leads'])
            ws.cell(row=idx, column=3, value=month['deals'])
            ws.cell(row=idx, column=4, value=month['revenue']).number_format = '€#,##0.00'
            ws.cell(row=idx, column=5, value=month['conversion']).number_format = '0.0%'

        # Line chart for trends
        chart = LineChart()
        chart.title = "Revenue Trend (Last 6 Months)"
        chart.x_axis.title = "Month"
        chart.y_axis.title = "Revenue (€)"
        chart.style = 13

        data = Reference(ws, min_col=4, min_row=3, max_row=len(monthly_data)+3)
        cats = Reference(ws, min_col=1, min_row=4, max_row=len(monthly_data)+3)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, "G3")

        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18

    def save(self, filename: str):
        """Save workbook to file."""
        self.wb.save(filename)
        return filename


def generate_sample_data():
    """Generate sample CRM data for demonstration."""
    brokers = [
        {'name': 'Michael Schmidt', 'leads': 45, 'deals': 8, 'revenue': 120000, 'conversion': 0.178},
        {'name': 'Anna Müller', 'leads': 52, 'deals': 11, 'revenue': 165000, 'conversion': 0.212},
        {'name': 'Thomas Weber', 'leads': 38, 'deals': 5, 'revenue': 75000, 'conversion': 0.132},
        {'name': 'Sarah Klein', 'leads': 61, 'deals': 14, 'revenue': 210000, 'conversion': 0.230},
        {'name': 'Daniel Becker', 'leads': 29, 'deals': 3, 'revenue': 45000, 'conversion': 0.103},
    ]

    months = []
    for i in range(6):
        month_date = date.today() - timedelta(days=30*i)
        months.insert(0, {
            'month': month_date.strftime('%B %Y'),
            'leads': random.randint(180, 250),
            'deals': random.randint(35, 55),
            'revenue': random.randint(500000, 700000),
            'conversion': random.uniform(0.15, 0.25)
        })

    return brokers, months


def main():
    """Generate complete w&k Connect CRM report."""
    print("📊 Generating w&k Connect CRM Report...")

    # Generate sample data
    broker_data, monthly_data = generate_sample_data()

    # Create report
    generator = CRMReportGenerator()
    generator.create_broker_performance_sheet(broker_data)
    generator.create_monthly_trend_sheet(monthly_data)

    # Save
    filename = f'wk_crm_report_{date.today().isoformat()}.xlsx'
    output_path = generator.save(filename)

    # Summary
    total_revenue = sum(b['revenue'] for b in broker_data)
    total_deals = sum(b['deals'] for b in broker_data)
    avg_conversion = sum(b['conversion'] for b in broker_data) / len(broker_data)

    print(f"\n✅ Report generated successfully!")
    print(f"📄 Location: {output_path}")
    print(f"\n📈 Summary:")
    print(f"   • Total Revenue: €{total_revenue:,.2f}")
    print(f"   • Total Deals: {total_deals}")
    print(f"   • Avg Conversion: {avg_conversion:.1%}")
    print(f"   • Worksheets: 2 (Broker Performance + Monthly Trends)")
    print(f"   • Charts: 2 (Bar Chart + Line Chart)")

    return output_path


if __name__ == '__main__':
    main()
